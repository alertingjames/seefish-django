import os
import string
import urllib

import io
import re
import requests

# from _mysql_exceptions import DataError, IntegrityError
from django.template import RequestContext

import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText

from django.core.files.storage import FileSystemStorage
from django.http import HttpResponse, JsonResponse, HttpResponseRedirect
from django.utils.datastructures import MultiValueDictKeyError

from rest_framework.permissions import AllowAny
from time import gmtime, strftime
import json
from django.contrib.sessions.backends.db import SessionStore
from django.contrib.sessions.models import Session
from django.conf import settings
from django import forms

import sys
from django.core.cache import cache
import mechanicalsoup
from bs4 import BeautifulSoup as soup
from urllib.request import Request, urlopen


def snowflakecustomerscrapping(request):
    base_url = "https://www.snowflake.com/customers/"
    dataList = []
    for i in range(1, 22):
        url = base_url + "?sf_paged=" + str(i)
        req = Request(url, headers={'User-Agent': 'XYZ/3.0'})
        webpage = urlopen(req, timeout=10).read()
        page_soup = soup(webpage, 'html.parser')
        results = page_soup.find('div', class_='search-filter-results-grid')
        elems = results.find_all('div', class_='cell')
        for elem in elems:
            title_elem = elem.find('img', class_='cs-logo')["alt"]

            site_link = 'www.' + title_elem.replace(' ','').lower() + '.com'

            data = {
                'company': title_elem,
                'website': site_link
            }
            dataList.append(data)

    import openpyxl
    from openpyxl.utils import get_column_letter

    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename=' + 'Snowflake_customers.xlsx'
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Customers"

    row_num = 0

    columns = [
        (u"Company Name", 30),
        (u"Company Website", 100),
    ]

    my_color = openpyxl.styles.colors.Color(rgb='FF99EDFC')
    my_fill = openpyxl.styles.fills.PatternFill(patternType='solid', fgColor=my_color)

    from openpyxl.styles import Alignment

    for col_num in range(len(columns)):
        c = ws.cell(row=row_num + 1, column=col_num + 1)
        c.value = columns[col_num][0]
        # set column width
        ws.column_dimensions[get_column_letter(col_num + 1)].width = columns[col_num][1]
        c.fill = my_fill
        c.alignment = Alignment(horizontal='center')

    row_num = 1
    for data in dataList:
        row_num = row_num + 1

        c = ws.cell(row=row_num, column=1)
        c.value = data['company']

        c = ws.cell(row=row_num, column=2)
        c.value = data['website']

    wb.save(response)
    return response



def getwebsite(company_name):
    url = "https://companyurlfinder.com/cuf"
    querystring = {"companyName":company_name,"api_key":"nVBdtHYTP70B3YUQa0IFlAutlQcwq34LsGTO88S6"}

    payload = ""
    headers = {
    'content-type': "application/json"
    }
    response = requests.request("GET", url, data=payload, headers=headers, params=querystring)
    return response.text


def saptest(request):
    dataList = []
    for i in range(1, 80):
        # page_num = request.GET['page']
        page_num = i
        url = 'https://www.sap.com/bin/sapdx/solrsearch'
        params = dict(
            showEmptyTags='false',
            isDateRange='false',
            isEventPeriod='false',
            isResourceCenter='false',
            highlighting='false',
            fuzzySearch='false',
            hideFacets='false',
            additionalProcess='false',
            showEventInfo='true',
            originalPath='/content/sapdx/countries/en_us/about/customer-stories',
            targetPath='/content/sapdx/languages/en_us/about/customer-stories',
            isFullTextSearch='false',
            pageLocale='en_us',
            json='{"componentPath":"/content/sapdx/languages/en_us/about/customer-stories/jcr:content/par/section_copy/section-par/resourcecenterdynamic/item_1594287740507","search":[],"pagePath":"/content/sapdx/languages/en_us/about/customer-stories","page":' + str(page_num) + ',"pageCount":40,"searchText":"","sortName":"lastModified","sortType":"desc"}',
        )
        resp = requests.get(url=url, params=params)
        resp = resp.json()
        results = resp.get('results')
        # return HttpResponse(json.dumps(results))
        if len(results) == 0: continue
        for res in results:
            title = ''
            business = ''
            try:
                title = res['title']
                if ':' in title:
                    business = title.split(':')[0]
            except KeyError:
                print('No key')

            description = ''
            try:
                description = res['description']
            except KeyError:
                print('No key')

            akamaiUrl = ''
            try:
                akamaiUrl = res['akamaiUrl'][0]
            except KeyError:
                print('No key')

            keywords = ''
            try:
                keywords = res['damSearchKeywords']
            except KeyError:
                print('No key')

            url = ''
            try:
                url = 'https:' + res['url']
            except KeyError:
                print('No key')

            data = {
                'title': title,
                'business': business,
                'description': description,
                'akamaiUrl': akamaiUrl,
                'keywords': keywords,
                'url': url,
            }
            dataList.append(data)

    # return HttpResponse(json.dumps(dataList))

    import openpyxl
    from openpyxl.utils import get_column_letter

    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename=' + 'SAP_customers.xlsx'
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Customers"

    row_num = 0

    columns = [
        (u"Business", 30),
        (u"Url", 80),
        (u"AkamaiUrl", 50),
        (u"Keywords", 50),
        (u"Title", 50),
        (u"Description", 100),
    ]

    my_color = openpyxl.styles.colors.Color(rgb='FF99EDFC')
    my_fill = openpyxl.styles.fills.PatternFill(patternType='solid', fgColor=my_color)

    from openpyxl.styles import Alignment

    for col_num in range(len(columns)):
        c = ws.cell(row=row_num + 1, column=col_num + 1)
        c.value = columns[col_num][0]
        # set column width
        ws.column_dimensions[get_column_letter(col_num + 1)].width = columns[col_num][1]
        c.fill = my_fill
        c.alignment = Alignment(horizontal='center')

    row_num = 1
    for data in dataList:
        row_num = row_num + 1

        c = ws.cell(row=row_num, column=1)
        c.value = data['business']

        c = ws.cell(row=row_num, column=2)
        c.value = data['url']

        c = ws.cell(row=row_num, column=3)
        c.value = data['akamaiUrl']

        c = ws.cell(row=row_num, column=4)
        c.value = data['keywords']

        c = ws.cell(row=row_num, column=5)
        c.value = data['title']

        c = ws.cell(row=row_num, column=6)
        c.value = data['description']

    wb.save(response)
    return response




def googlecustomerstest(request):
    dataList = []
    resp = requests.get('https://cloud.google.com/customers/customers.json?hl=en')
    resp = resp.json()
    results = resp.get('customers')
    # return HttpResponse(json.dumps(results))
    if len(results) > 0:
        for res in results:
            name = ''
            blurb = ''
            videolink = ''
            bloglink = ''
            casestudylink = ''
            logo = ''
            products = ''
            industries = ''
            regions = ''

            try:
                name = res['name']
            except KeyError:
                print('No key')

            try:
                blurb = res['blurb']
            except KeyError:
                print('No key')

            try:
                videolink = res['videolink']
            except KeyError:
                pass
            except io.BlockingIOError:
                pass

            try:
                bloglink = res['bloglink']
            except KeyError:
                pass
            except io.BlockingIOError:
                pass

            try:
                casestudylink = res['casestudylink']
                if casestudylink != '' and not 'http' in casestudylink:
                    casestudylink = 'https://cloud.google.com' + casestudylink
            except KeyError:
                print('No key')

            try:
                logo = 'https://cloud.google.com/images/customers/directory/' + res['logo']
            except KeyError:
                print('No key')

            try:
                products = res['products']
                products = str(products).replace('[','').replace(']','').replace('\'','')
            except KeyError:
                print('No key')

            try:
                industries = res['industries']
                industries = str(industries).replace('[','').replace(']','').replace('\'','')
            except KeyError:
                print('No key')

            try:
                regions = res['regions']
                regions = str(regions).replace('[','').replace(']','').replace('\'','')
            except KeyError:
                print('No key')

            data = {
                'name': name,
                'blurb': blurb,
                'videolink': videolink,
                'bloglink': bloglink,
                'casestudylink': casestudylink,
                'logo': logo,
                'products': products,
                'industries': industries,
                'regions': regions,
            }
            dataList.append(data)


    # return HttpResponse(json.dumps(dataList))

    import openpyxl
    from openpyxl.utils import get_column_letter

    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename=' + 'google_customers.xlsx'
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Customers"

    row_num = 0

    columns = [
        (u"Name", 30),
        (u"Logo", 50),
        (u"Blurb", 80),
        (u"Casestudy Link", 50),
        (u"Video Link", 50),
        (u"Blog Link", 50),
        (u"Products", 50),
        (u"Industries", 50),
        (u"Regions", 50),
    ]

    my_color = openpyxl.styles.colors.Color(rgb='FF99EDFC')
    my_fill = openpyxl.styles.fills.PatternFill(patternType='solid', fgColor=my_color)

    from openpyxl.styles import Alignment

    for col_num in range(len(columns)):
        c = ws.cell(row=row_num + 1, column=col_num + 1)
        c.value = columns[col_num][0]
        # set column width
        ws.column_dimensions[get_column_letter(col_num + 1)].width = columns[col_num][1]
        c.fill = my_fill
        c.alignment = Alignment(horizontal='center')

    row_num = 1
    for data in dataList:
        row_num = row_num + 1

        c = ws.cell(row=row_num, column=1)
        c.value = data['name']

        c = ws.cell(row=row_num, column=2)
        c.value = data['logo']

        c = ws.cell(row=row_num, column=3)
        c.value = data['blurb']

        c = ws.cell(row=row_num, column=4)
        c.value = data['casestudylink']

        c = ws.cell(row=row_num, column=5)
        c.value = data['videolink']

        c = ws.cell(row=row_num, column=6)
        c.value = data['bloglink']

        c = ws.cell(row=row_num, column=7)
        c.value = data['products']

        c = ws.cell(row=row_num, column=8)
        c.value = data['industries']

        c = ws.cell(row=row_num, column=9)
        c.value = data['regions']

    wb.save(response)
    return response




def indeedjobs(request):
    url = 'https://www.indeed.com/jobs?q=hiring+immediately&l=New+York%2C+NY&from=homepage_relatedQuery'
    browser = mechanicalsoup.StatefulBrowser()
    browser.open(url)

    return HttpResponse(browser.url)


























